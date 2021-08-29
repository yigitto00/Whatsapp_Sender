import pywhatkit as kit                     #xlsx dosyasını okumak için gerekli kütüphane.
from openpyxl import Workbook,load_workbook #Watsapp dan mesaj göndermek için gerekli kütüphane.

print("""
______________________________
|       ByGoldTeam           |
|       DeveloperByYigitto0  |
|       ChallengeByCan Deger |
|____________________________| 

""")

print("Not: Dkikada sadece 1 mesaj gönderebilirsiniz.")
wb = load_workbook("numbers.xlsx") #xlsx dosyasını açıyoruz.
ws = wb.sheetnames


ws = wb['telefon_number']# Bölüm ü seçiyoruz.


#Gerekli sorular.
saat = input("Saat(Dijital):")
dakika = input("Dakika:")
mesaj = input("Mesaj:")
print("""
Not: eğer 1 mesaj gönderecekseniz 2, 2 mesaj ise 3 vb.
""")
numarakaç = input("Kaç numaraya mesaj göndereyim:")

#Değerlerimizi intager ve ya string e çeviriyoruz.
saat = int(saat)
dakika = int(dakika)
numarakaç = int(numarakaç) 




#Gönderme işlemi.
for satir in range(1,numarakaç):
    for sutun in range(1,2):        
        numara = "+" + str(ws.cell(satir,sutun).value)
        kit.sendwhatmsg(numara, mesaj, saat,dakika)
        dakika = dakika+1


